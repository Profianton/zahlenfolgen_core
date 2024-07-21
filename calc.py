# Importiert selbstgeschriebene Funktion zum Ausrechnen von Formeln
from rechner.math_eval import calculate as calc
from time import time
import os
from fractions import Fraction
from openpyxl import Workbook, drawing, styles
from matplotlib import pyplot as plt
plt.switch_backend('agg')                    # stabilisiert Plots
Font = styles.Font


def stringify_number(number):
    """Funktion, um Zahlen schön als String darstellen. Damit können Zahlen leicht lesbar ausgegeben werden. z.B. als 1+1/2


    Args:
        number (int|float): number to be stringified

    Returns:
        string: number as a readable string
    """
    fraction = Fraction(number).limit_denominator()
    whole_part = int(fraction)
    fractional_part = fraction - whole_part
    if fractional_part == 0:
        return str(whole_part)
    elif whole_part == 0:
        return str(fractional_part)
    else:
        return f"{whole_part}+{fractional_part}"


def check_expression(string: str):
    """Funktion,um zu Überprüfen, ob eine nicht-rekursive Folge die Richtige ist

     Args:
         string (string): formel

     Returns:
         tuple[string, float|in, Literal['Normal']] | None: _description_
     """
    # if "n" not in string:
    #     return None
    # print(string, end=" "*(20-len(string)))
    for idx, num in enumerate(list_with_numbers):
        """try:
            print(f"f({idx+1})=",
                    stringify_number(calculate(idx + 1, string)), end="\t")
        except:
            print(f"f({idx+1})=NaN", end="")"""
        if calculate(idx + 1, string) != num:  # Wenn eine Folge nicht passt, gebe None zurück
            # print()
            return None
    # Rechne nächstes Element aus, wenn die ersten Folgeglieder passen
    n = len(list_with_numbers) + 1
    result = calculate(n, string)
    if result != None:
        return string, result, "Normal"


def normal_move(complexity_left, string):
    """Einen Operator und eine Zahl ergänzen, wenn Komplexität übrig ist, sonst Formel prüfen
    Die Komplexität wird erhöht, wenn ein Operator, wie +, -, *, /, ^ eingefügt wird.

    Args:
        complexity_left (int): Anzahl der Rechenzeichen, die noch an die Formel angehängt werden sollen. D.h. Anzahl der Rechenzeichen, die noch übrig sind.
        string (_type_): Bisherige Formel

    Returns:
        tuple[str, int | float, Literal['Normal']] | None: Formel mit nächstem Element und "normal", wenn die Formel funktioniert, sonst None
    """
    if complexity_left == 0:
        # Formel prüfen
        return check_expression(string)
    # operator anhängen
    for operator in operators:
        if operator["c"] <= complexity_left:
            # Zahl anhängen
            for num in numbers:
                result = normal_move(
                    complexity_left -
                    operator["c"], string + operator["z"] + num
                )
                # Ergebnis der Formel zurückgeben, wenn die Formel passt
                if result != None:
                    return result


def calculate(n, string):
    """Funktion zum Ausrechnen von nicht-rekursiven Formeln

    Args:
        n (int): n von f(n)
        string (string): formel

    Returns:
        int|float|None: int oder float als Ergebnis, None bei Fehlern
    """

    string = string.replace("n", str(n))
    try:
        return calc(string)
    except:
        return None


def tokenize_fix(num):
    """Negative Zahlen für calc vorbereiten


    Args:
        num (int|float): eine eventuell negative Zahl, die für die Funktion calc vorbereitet werden muss 

    Returns:
        string: Zahl als String; negative Zahlen werden als 0-Zahl dargestellt
    """
    if num < 0:
        return f"(0-{-num})"
    else:
        return str(num)


def calculate_recursive(n, string):
    """Funktion zum Ausrechnen von rekursiven Formeln

    Args:
        n (int): n von f(n)
        string (string): Formel

    Returns:
        int|float|None: int oder float als Ergebnis, None bei Fehlern
    """
    # Elemente f(n-1) und f(n-2) durch den Wert ersetzen und n einfügen
    string = (
        string.replace("f(n-2)", tokenize_fix(list_with_numbers[n - 3]))
        .replace("f(n-1)", tokenize_fix(list_with_numbers[n - 2]))
        .replace("n", str(n))
    )
    try:
        # ausrechnen
        return calc(string)
    except ZeroDivisionError:
        return None


def recursive_normal_move(complexity_left, string):
    """Einen Operator und eine Zahl ergänzen, solange Komplexität übrig ist, ansonsten Formel prüfen für nicht-rekursive Folgen
Die Komplexität wird erhöht, wenn ein Operator, wie +, -, *, /, ^ eingefügt wird.

    Args:
        complexity_left (int): Anzahl der Rechenzeichen, die noch an die Formel angehängt werden sollen. D.h. Anzahl der Rechenzeichen, die noch übrig sind.
        string (_type_): Bisherige Formel

    Returns:
        tuple[str, int | float, Literal['Recursive']] | None: Formel mit nächstem Element und "Recursive", wenn die Formel funktioniert, sonst None
    """
    if complexity_left == 0:
        if not (("f(n-2)" in string) or ("f(n-1)" in string)):
            return None
        # print(string)
        items_to_check = list(enumerate(list_with_numbers))
        if "f(n-2)" in string:
            items_to_check = items_to_check[2:]
        else:
            items_to_check = items_to_check[1:]
        for idx, num in items_to_check:
            if calculate_recursive(idx + 1, string) == num:
                1
            else:
                return None
        i = len(list_with_numbers) + 1
        result = calculate_recursive(i, string)
        if result != None:
            return string, result, "Recursive"
        else:
            return None
    for operator in operators:
        if operator["c"] <= complexity_left:
            for num in numbers_recusive:
                result = recursive_normal_move(
                    complexity_left -
                    operator["c"], string + operator["z"] + num
                )
                if result != None:
                    return result


def recusive_first_move(complexity):
    """Funktion zum Hinzufügen der ersten Zahl und aufrufen von recursive_normal_move


    Args:
        complexity (int): maximale Komplexität, die die Formel haben darf

    Returns:
        tuple[string, int | float, Literal['Recursive']] | None: Formel, nächste Zahl und "recursive" bei einer gefundenen Formel, sonst None
    """
    global numbers_recusive
    numbers_recusive = numbers + ["f(n-1)", "f(n-2)", "(f(n-1)-f(n-2))"]
    for number in numbers_recusive:
        result = recursive_normal_move(complexity, number)
        if result != None:
            return result


def first_move(complexity):
    """Funktion zum Hinzufügen der ersten Zahl und Aufrufen von "normal_move" und recursive_fist_move, falls die Komplexität größer als 1 ist.

    Args:
        complexity (int): maximale Komplexität, die die Formel haben darf

    Returns:
        tuple[string, int | float, Literal['Normal']] | None: Formel, nächste Zahl und "Normal" bei einer gefundenen Formel, sonst None
    """
    for num in numbers:
        result = normal_move(complexity, num)
        if result != None:
            return result
    if complexity - 1 >= 0:
        return recusive_first_move(complexity - 1)


def run_with_complexity(max_complexity):
    """Wrapper (Übersichtliches Interface) für "first_move"

    Args:
        max_complexity (int): maximale Komplexität der Formeln

    Returns:
        tuple[str, int | float, Literal['Normal'] | Literal['Recursive']] | None: Formel oder None
    """
    global operators, numbers
    # Erlaubte Operatoren festlegen
    operators = [
        {"z": "*", "c": 1},
        {"z": "/", "c": 1},
        {"z": "+", "c": 1},
        {"z": "-", "c": 1},
        {"z": "^", "c": 1},
    ]
    # Erlaubte Zahlen festlegen
    numbers = [f"{num}" for num in list(
        range(1, 6))] + ["n", "(n-1)", "(0-1)", "10"]

    return first_move(max_complexity)


def get_inputs():
    """Funktion zum Erhalten von Nutzereingaben, wenn nicht das GUI, sondern die Kommandozeile verwendet wird

    Returns:
        list[float | int]: nutzereingaben
    """
    list_with_numbers = []
    os.system("cls" if os.name == "nt" else "clear")
    while True:
        if list_with_numbers == []:
            string = f"Enter a number:"
        else:
            string = f"Enter a number (or press 'Enter' to finish) \n {', '.join(
                [stringify_number(number) for number in list_with_numbers])}, "
        number = input(string)
        os.system("cls" if os.name == "nt" else "clear")
        if number == "" and list_with_numbers != []:
            break
        else:
            try:
                number = eval(number)
                if type(number) == float or type(number) == int:
                    list_with_numbers.append(number)
            except:
                print("Invalid input")
    print(
        f"\nYou entered the following numbers: {
            ', '.join([stringify_number(number) for number in list_with_numbers])}"
    )
    return list_with_numbers


real_print = print   # Die standard-python print-Funktion habe ich umbenannt


def solve(numbers) -> tuple[str, list[float | int], str, str, list[float | int]]:
    """Zentrale Funktion zum Lösen von Formeln

    Args:
        numbers (list[int | float]): zahlenfolge zum weiterführen

    Returns:
        tuple[str, list[float | int], Literal['Normal', 'Recursive'], str, list[float | int]]: Formel, Weitergeführte Zahlen, Typ der Formel, Log, Reihe
    """
    plt.clf()   # Plots Löschen
    global log
    log = ""

    def print(x):
        """Eigene, customized print-Funktion, um alles mit zu loggen und an das Frontend zu schicken"""
        global real_print
        real_print(x)
        global log
        log += str(x) + "\n"

    print(f"solving {numbers}")
    global list_with_numbers
    list_with_numbers = numbers
    global complexity
    complexity = 0
    solved = False
    start_time = time()
    while not solved:
        print(f"Running with complexity {complexity}.")
        complexity_start_time = time()
        output = run_with_complexity(complexity)
        print(
            f"complexity {complexity} finished in {
                round(time()-complexity_start_time, 5)} seconds."
        )
        if output != None:
            formula, next_number, type = output

            if type == "Normal":
                print(
                    f"the formula was '{formula}' and the next numbers are {', '.join([stringify_number(
                        calculate(i, formula)) for i in range(len(list_with_numbers)+1, len(list_with_numbers)+5)])}."
                )
                # Plotten von nicht-rekursiven Folgen
                plt.plot(
                    [i /
                        10 for i in range(0, (len(list_with_numbers) + 10) * 10-9, 1)],
                    [
                        calculate(i / 10, formula)
                        for i in range(0, (len(list_with_numbers) + 10) * 10-9, 1)
                    ],
                )
                plt.plot(

                    [
                        calculate(i, formula)
                        for i in range(0, (len(list_with_numbers) + 10), 1)
                    ],
                    "o "
                )
                plt.xticks(range(0, len(list_with_numbers)+10, 1))
            else:
                ols_list_with_numbers = list_with_numbers.copy()
                for i in range(len(list_with_numbers) + 1, len(list_with_numbers) + 10):
                    try:
                        list_with_numbers.append(
                            calculate_recursive(i, formula))
                    except:
                        break
                # Plotten von rekursiven Folgen
                plt.plot(
                    list_with_numbers,
                    "o-"
                )
                plt.xticks(range(0, len(list_with_numbers), 1))
                print(
                    f"the recursive formula was 'f(n)={formula}' and the next numbers are {', '.join(
                        [stringify_number(number) for number in list_with_numbers[len(ols_list_with_numbers):]])}."
                )

            print(
                f"solved complexity {complexity} problem in {
                    round(time()-start_time, 5)} seconds."
            )
            # Plots generieren
            # Plot Folge
            plt.title("f(n)")

            plt.savefig(f"{zahlenfolgen_dir}/plot.png")

            # Plot Reihe
            plt.clf()
            for i in range(len(list_with_numbers) + 1, len(list_with_numbers) + 10):
                list_with_numbers.append(calculate_recursive(i, formula))
            plt.xticks(range(0, len(list_with_numbers), 1))
            plt.title("∑f(n)")
            series = [sum(list_with_numbers[:i+1]) for i in range(
                0, len(list_with_numbers), 1)]  # Mathematische Reihe
            plt.plot(
                series,
                "o-"
            )

            plt.savefig(f"{zahlenfolgen_dir}/series.png")
            gen_excel(formula, list_with_numbers, type)
            solved = True
            print((formula, list_with_numbers, type))
            return (formula, list_with_numbers, type, log, series)

        else:
            complexity += 1


def gen_excel(formula, list_with_numbers, type):
    """Funktion, um Excel-Datei zu generieren"""
    wb = Workbook()
    ws = wb.active
    ws["A1"].value = "n"
    ws['A1'].font = Font(bold=True)
    ws["B1"].value = "Folge f(n)"
    ws['B1'].font = Font(bold=True)
    ws["C1"].value = "Reihe ∑f(n)"
    ws['C1'].font = Font(bold=True)
    for n in range(1, 10+1, 1):
        ws[f"A{n+1}"].value = n
        ws[f"B{n+1}"].value = \
            "="+(formula.replace("f(n-1)", f"B{n}")
                 .replace("f(n-2)", f"B{n-1}")
                 .replace("n", f"A{n+1}"))
        ws[f"C{n+1}"].value = f"=C{n}+B{n+1}"
    ws["C2"].value = "=B2"

    if type == "Recursive":
        ws["B2"] = list_with_numbers[0]
        if "f(n-2)" in formula:
            ws["B3"] = list_with_numbers[1]

    ws["E1"].value = "Formel"
    ws['E1'].font = Font(bold=True)
    ws["E2"].value = f"f(n)={formula}"
    ws['E2'].font = Font(bold=True)

    if type == "Recursive":
        ws["E3"] = f"f(1)={list_with_numbers[0]}"
        ws['E3'].font = Font(bold=True)

        if "f(n-2)" in formula:
            ws["E4"] = f"f(1)={list_with_numbers[1]}"
            ws['E4'].font = Font(bold=True)

    img = drawing.image.Image(f"{zahlenfolgen_dir}/plot.png")
    img.anchor = 'G1'
    ws.add_image(img)
    img = drawing.image.Image(f"{zahlenfolgen_dir}/series.png")
    img.anchor = 'P1'
    ws.add_image(img)

    wb.save(f"{zahlenfolgen_dir}/excel.xlsx")


zahlenfolgen_dir = "/".join(__file__.replace("\\", "/").split("/")[:-2])

# Referenzcode zum Testen der Performanz des Programms
if __name__ == "__main__":
    start = time()
    for _ in range(10):
        solve([1, 1, 2, 3, 5, 8, 13])
    print(f"time_average: {(time()-start)/10}")
